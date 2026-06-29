"""Phase 1 の最小テスト。`python tests/test_basic.py` で実行（pytest 不要）。

※ 値はすべて架空（合成）。実患者データは使わない。
"""
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import load_workbook
from kenshin import compute, excel_writer, pipeline


def test_compute():
    assert compute.age_at("S50.01.01", "2026-06-25") == 51
    assert compute.age_at("S55.05.05", "2026-06-25") == 46
    # 受診日が誕生日の前 → 1 歳若い
    assert compute.age_at("S55.12.31", "2026-06-25") == 45
    assert compute.parse_wareki("H1.01.08").year == 1989
    assert compute.bmi(170.0, 65.0) == 22.5
    assert compute.bmi(160.0, 60.0) == 23.4
    assert compute.fmt_birth_with_age("S50.01.01", "2026-06-25") == "S50.01.01 (51歳)"


def test_record_build():
    rec = pipeline.record_from_values(
        {"身長": 170.0, "体重": 65.0, "腹囲": 80.0,
         "生年月日raw": "S50.01.01", "受診日raw": "2026-06-25"}
    )
    assert rec.get("身長体重").value == "170.0 / 65.0"
    assert rec.get("BMI腹囲").value == "22.5 / 80.0"
    assert rec.get("生年月日").value == "S50.01.01 (51歳)"


def test_writer_flags_uncertain():
    config = excel_writer.load_config(ROOT / "configs" / "企業.json")
    values = {
        "氏名": "サンプル太郎",
        "GOT": "20",
        "自覚症状": {"value": "x", "needs_check": True},
        "血色素量": {"value": "14.0", "confidence": 0.3},
    }
    with tempfile.TemporaryDirectory() as d:
        out = Path(d) / "out.xlsx"
        flagged = pipeline.run_from_values(values, config, out, base_dir=ROOT)
        keys = {k for k, _ in flagged}
        assert keys == {"自覚症状", "血色素量"}
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == "個人票"
        assert ws["H5"].value == "20"            # 書き込まれた
        assert ws["H3"].value is None            # 要確認 → 空欄
        assert ws["H3"].fill.fgColor.rgb == "00FFFF00"  # 黄色
        assert ws["B16"].value is None
        assert ws["E16"].value == "12.1-14.5"    # 基準範囲は保持（静的・患者値ではない）
        assert wb.sheetnames == ["個人票"]        # 記入例シートは出力に含めない


if __name__ == "__main__":
    test_compute()
    test_record_build()
    test_writer_flags_uncertain()
    print("OK: all tests passed")
