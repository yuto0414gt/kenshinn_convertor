"""個人票テンプレ Excel へ実測値を書き込む。

方針(合意済み):
- テンプレの『テンプレ』シートを複製し、そこへ実測値を書き込む。体裁・基準範囲・凡例は壊さない。
- 確信度が低い / 読めなかった項目は、値を入れず**セルを黄色く塗る**（医師がそこだけ手で埋める）。
- 判定 A〜E・生活アドバイス・総合所見は Phase 1 では触らない（空欄のまま）。
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .models import PatientRecord

# 要確認セルの色（黄色）。openpyxl は ARGB 8 桁。
FLAG_FILL = PatternFill("solid", fgColor="FFFF00")


def load_config(path: str | Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def write(record: PatientRecord, config: dict, out_path: str | Path,
          base_dir: str | Path = ".", conf_threshold: float = 0.6) -> list[tuple[str, str]]:
    """record を config の対応づけに従ってテンプレへ書き込み、out_path へ保存。

    戻り値: 要確認として色付けした (field_key, セル) のリスト。
    """
    base_dir = Path(base_dir)
    template_path = base_dir / config["template"]
    wb = load_workbook(template_path)
    ws = wb[config["template_sheet"]]

    flagged: list[tuple[str, str]] = []
    for key, cell in config["fields"].items():
        fv = record.get(key)
        if fv is None:
            # その項目自体を渡していない → テンプレの空欄のまま（色も付けない）
            continue
        target = ws[cell]
        if fv.is_uncertain(conf_threshold):
            target.value = None
            target.fill = FLAG_FILL  # 罫線は別属性なので保持される
            flagged.append((key, cell))
        else:
            target.value = fv.value

    # 出力シート名へリネーム（テンプレ原本と区別）。検証用『記入例』シートは出力から除く。
    ws.title = config.get("output_sheet", ws.title)
    for name in list(wb.sheetnames):
        if name != ws.title:
            del wb[name]

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return flagged
